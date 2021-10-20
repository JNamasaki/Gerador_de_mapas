import folium
import pandas as pd
import webbrowser
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl as pyxl
import time
import sys

data = 'Endereços.xlsx'

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
navegador = webdriver.Chrome(executable_path= ChromeDriverManager().install(),options=options)
navegador.get('https://www.google.com.br/maps/')

end = pd.read_excel(data)
wb = pyxl.load_workbook(data)

sheet = wb['Planilha1']
cell_sheet = sheet.cell(row=1,column=4,value='Long')
cell_sheet =sheet.cell(row=1,column=5,value='Lat')

for i in range(len(end)):
    valor_cell = end['Endereços'] [i]
    busca = navegador.find_element_by_xpath('//*[@id="searchboxinput"]').send_keys(valor_cell)
    time.sleep(2)
    busca = navegador.find_element_by_id("searchbox-searchbutton").click()
    time.sleep(2)

    coord = navegador.current_url

    busca = navegador.find_element_by_xpath('//*[@id="searchboxinput"]').clear()
    time.sleep(2)

    met = coord.split('/')
    if met[4] =='search':
        coor[0] = 'Error'
        coor[1] = 'Error'
    elif met[4] == 'place':
        coor = coord.split("@")[1]
        coor = coor.split(",",2)
    for j in range(len(coor)-1):
        if j == 0:
            d = sheet.cell(row=(i+2), column=4,value= coor[j])
        else :
            d = sheet.cell(row=(i+2), column=5,value= coor[j])

wb.save(data)

navegador.close()

# read the data from string
df = pd.read_excel(data)

mapa = folium.Map(location=[-21.53607263898476, -49.859271331350875],zoom_start=14)
folium.Marker(location=[-21.53607263898476, -49.859271331350875],
    popup='<i>Ponto Inical</i>',  icon=folium.Icon(color='green')
    ).add_to(mapa)
for i in range(len(df)):
    des = df['Long'][i]
    if (des != 'Error'):
        long = df['Long'][i]
        lat = df['Lat'][i]
        title = df['Endereços'][i]
        folium.Marker(location=[long , lat],
        popup='<i>'+title+'</i>', 
        ).add_to(mapa)
    else:
        continue
mapa.save('Mapa.html')

webbrowser.open_new_tab('Mapa.html')

sys.exit()